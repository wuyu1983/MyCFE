# -*- coding: utf-8 -*-
import requests
import os
import sqlite3
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import time
from pprint import pprint
from bs4 import BeautifulSoup as bs
from lxml import etree
from lxml import html

import gevent
from gevent.queue import Queue
from gevent.pool import Pool
from gevent import Greenlet
from gevent import monkey
monkey.patch_all()

class CFE():
    path = r"D:\others"  # 默认路径
    filename = 'cfebom5.sqlite'  # 默认数据库文件名
    bom_table = 'bom'  # 默认bom数据表名
    failure_report_table = 'failure_report' # 失败料号信息

    boms=Queue()
    success_report=[]
    failure_report=[]
    table_title = ["Part", "PartDescription", "Lv", "BOM_Quantity", "AltGroup", "AltPri", "AltPercentage",
                    "CostPlant", "CostName", "source", "Is_Keypart", "MATKL", "MATKL_Desc",
                    "WEIGHTED_AVERAGE", "LATTEST_PROCUR", "M1", "M2", "M3", "M4", "M5", "M6", "MFG_PN", "MFG_DESC",
                    "Product_Family", "Cycle", "Plant"]

    def cost_bom_prepare():
        # 准备好cookies，建立并返回Session
        s = requests.Session()
        # 获取server授权的JSESSIONID，准备好Session的headers，传递给查BOM的函数
        url = 'http://planning.lenovo.com/abpp_ui_cfe_midh/core/login.jsp?ACTIVITY=manage_enterprise&CONTENT_URL=&FROM=timeout'
        header = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:48.0) Gecko/20100101 Firefox/48.0',
              'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
              'Accept-Language': 'zh-CN,zh;q=0.8,en-US;q=0.5,en;q=0.3',
              'Accept-Encoding': 'gzip, deflate',
              'Connection': 'keep-alive',
              'Upgrade-Insecure-Requests': '1'
              }
        s.headers.update(header)
        r1 = s.get(url=url)     # first request

        url = 'http://planning.lenovo.com/abpp_ui_cfe_midh/core/login/controller/login.x2c'
        header['Referer'] = 'http://planning.lenovo.com/abpp_ui_cfe_midh/core/login.jsp?ACTIVITY=manage_enterprise&CONTENT_URL=&FROM=timeout'
        header['Cookies'] = 'JSESSIONID=' + r1.cookies.get_dict()['JSESSIONID']

        s.cookies.update(r1.cookies)
        s.headers.update(header)
        payload = {'CHANGE_PASSWD': "no",
               'FORGOT_PASSWD': "no",
               'CONTENT_URL': "",
               'SCR_WIDTH': "1158",
               'USER_NAME': "wuyu4",
               'PASSWORD': "furthermore_2013"}
        r2 = s.post(url=url, data=payload, allow_redirects=False)       # second request

        header['Cookies'] = 'JSESSIONID=' + r2.cookies.get_dict()['JSESSIONID']
        header['Referer'] = 'http://planning.lenovo.com/abpp_ui_cfe_midh/costBomM.action'
        s.headers.update(header)
        return s

    def fetch_bom(s,query_load):
        # 根据准备好的Session去CFE服务器获得响应，并解析为soup，返回soup

        url = 'http://planning.lenovo.com/abpp_ui_cfe_midh/costBomM.action'
        r = s.post(url=url, data=query_load)
        soup = bs(r.content, 'lxml')
        if (r.content == b"No records found."):
            soup = "No records found."

            #failure_report.put_nowait(query_load)
            return soup

        return soup

    def build_bom(soup,pn):
        # 根据读取的soup，分析并抓取数据，并将数据构造为一个二维列表bom
        bom = []
        # uprows是cost bom的上半部分，包括各种parts和semi
        try:
            uprows = soup.select('div')[0].select('tr')[1:]
        except:
            print(pn,'error in uprows')
            #print(soup.prettify())
            return bom
        num_of_uprows = len(uprows)  # 第一个tr标签用来显示part, description这些

        # downrows是cost bom的下半部分，包括BOME和各种adder，以及economic cost等
        try:
            downrows = soup.select('div')[1].select('tr')[1:]  # 已剔除标题栏的tr，只保留数据部分
        except:
            print(pn,'error in downrows')
            #print(soup.prettify())
            bom = []
            return bom
        num_of_downrows = len(downrows)

        should_be_int = [2]     # 指出类型为int的列号
        should_be_float = [3, 6] + list(range(13, 21))      # 指出类型为float的列号
        for row in range(num_of_uprows):
            line = []
            for i in range(21):
                cell = uprows[row].select('td')[i].get_text().replace('\xa0', '').replace(' ', '')
                # 将部分字段内容强制转换成整数或浮点，以方便后面写入SQL，其他字段维持str不变
                if (i in should_be_int):
                    cell = int(cell)
                elif (i in should_be_float):
                    cell = float(cell)
                line.append(cell)
            bom.append(line)

        for row in range(num_of_downrows):
            line = []
            line.append(downrows[row].select('td')[0].get_text().replace('\xa0', ''))
            line += [''] * 12
            for i in range(8):
                cell = float(downrows[row].select('td')[i + 1].get_text().replace('\xa0', ''))
                line.append(cell)

            bom.append(line)
        # print(pn, ' in build_bom len(bom) is ',len(bom))
        return bom

    def search_in_cfe(pn, pn_desc, productfamily, cycle, plant):
        # 从CFE中获得相关BOM信息，并返回bom_array
        # 构造一个search query load

        s = CFE.cost_bom_prepare()

        # 准备好JSESSIONID

        start_time=time.time()
        print('Start to query {}...'.format(pn))
        query_load = {'method': 'search',
                      'cycleId': cycle,
                      'brandId': 'PHONE',
                      'plantId': plant,
                    'productFamilyId': productfamily,
                    'subgeoId': 'undefined',
                    'countryId': pn[-2:],
                    'assemblyId': pn,
                    'isExpandAll': '1'}

        soup = CFE.fetch_bom(s=s,query_load=query_load)         # 从CFE获得响应，转换为soup
        print('Constructing soup of {}'.format(pn))
        if (soup == "No records found."):       # BOM doesn't exist at all
            print("For {},{},{},{},no records found.".format(cycle, pn, plant, productfamily))
            CFE.failure_report.append([pn,pn_desc,productfamily,cycle,plant])

            bom_array = []
            end_time=time.time()
            print('{} bom is not existed... it takes {} '.format(pn, end_time - start_time))
            return bom_array
        else:
            bom_array = CFE.build_bom(soup, pn)

            # print(pn, ' in search_in_cfe len(bom) is ',len(bom_array),'and len(boms) is',len(CFE.boms))
        if len(bom_array) == 0:         # BOM has error
            print("For {},{},{},{},no records found.".format(cycle, pn, plant, productfamily))
            CFE.failure_report.append([pn, pn_desc, productfamily, cycle, plant])
            bom_array = []
            end_time = time.time()
            print('{} bom is not existed... it takes {} '.format(pn, end_time - start_time))
            return bom_array
        else:
            for line in bom_array:
                line+=[pn,pn_desc,productfamily,cycle,plant]
            CFE.success_report.append([pn, pn_desc, productfamily, cycle, plant])
            CFE.boms.put_nowait(bom_array)

        s.close()

        end_time = time.time()
        print('{} bom is ready... it takes {} '.format(pn,end_time-start_time))
        return bom_array

    def import_mfg():
        path = r"D:\1. Cost Tape\16年10月\Model list"
        filename = 'Sales Model List for FQ4 M0 Tape Release V2.xlsx'
        wb=openpyxl.load_workbook(os.path.join(path,filename))
        ws=wb.get_sheet_by_name('wuyu')

        mfgs=[]
        for row in range(3,1069):
            print("Loading No.{} mfg...".format(row-2))
            mfg=['','','','','']      #"MFG_PN","MFG_DESC","Product_Family","Cycle","Plant"
            mfg[0]=str(ws.cell(row=row,column=1).value)             # MFG_PN
            mfg[1]=str(ws.cell(row=row,column=2).value)             # MFG_DESC
            mfg[2]=str(ws.cell(row=row,column=4).value).upper()     # Product_Family
            mfg[3]="201610FINAL_B"                                # Cycle
            mfg[4]=str(ws.cell(row=row,column=17).value).upper()     # Plant
            if (mfg[4]=='LNV WH'):
                mfg[4]='6161'
            elif (mfg[4]=='LNV XM'):
                mfg[4]='6165'
            elif (mfg[4]=='BYD - Wuhan'):
                mfg[4]='6161'

            # elif (not mfg[4] in ['LONGCHEER','HUAQIN','ONTIM','WINGTECH','HIPAD','TDK']):
            #    mfg[4]='6161'
            # else:
            #    mfg[4]='6161'

            mfgs.append(mfg)

        return mfgs


    def db_init():
        # 设置并连接数据库
        connection = sqlite3.connect(os.path.join(CFE.path, CFE.filename))
        c = connection.cursor()
        # 若数据库不存在，则创建一个bom数据库和数据表
        # 该数据表将包括MFG的PN，MFG描述
        column_names = CFE.table_title

        int_columns = ["Lv"]
        float_columns = ["BOM_Quantity", "AltPercentage", "WEIGHTED_AVERAGE", "LATTEST_PROCUR", "M1", "M2", "M3", "M4",
                         "M5", "M6"]

        # 制作SQL创建表的指令
        sql = ""
        for i in range(len(column_names)):
            if (column_names[i] in int_columns):
                column_type = ' INT'
            elif (column_names[i] in float_columns):
                column_type = ' FLOAT'
            else:
                column_type = ' VARCHAR(100)'
            column_names[i] += column_type
        sql = ','.join(column_names)

        try:
            c.execute("CREATE TABLE {}({})".format(CFE.bom_table, sql))
        except:
            pass

        # 创建失效报告表
        try:
            c.execute('CREATE TABLE {}(MFG_PN VARCHAR(100), MFG_DESC VARCHAR(100),Product_Family VARCHAR(100),Cycle VARCHAR(100),Plant VARCHAR(100))'.format(CFE.failure_report_table))
        except:
            pass

        connection.commit()
        connection.close()

        return


    def bom_into_db():
        # 将bom数据，包括MFG料号和描述，以及各种parts、bome等写入数据库

        conn = sqlite3.connect(os.path.join(CFE.path, CFE.filename))
        c = conn.cursor()
        new_boms=[]
        # print('in bom_into_db len(boms) is',len(CFE.boms))
        while not CFE.boms.empty():         # gevent.queue.Queue cannot be used as iterator, it must be converted into a list
            new_boms.append(CFE.boms.get_nowait())
        print('new boms lenth ',len(new_boms))
        for bom in new_boms[:-1]:
            # 先将数据库中原有记录删除
            #print(bom[0])
            #print(bom[1])
            c.execute("DELETE FROM {} WHERE MFG_PN='{}' AND Cycle='{}' AND Plant='{}'".format(CFE.bom_table,bom[0][-5],bom[0][-2],bom[0][-1]))
            line = []
            place_holder = ','.join('?' * (len(bom[0])))  # make a {?,?,?,...,?}
            for row in bom:
                c.execute("INSERT INTO {} VALUES ({})".format(CFE.bom_table, place_holder), row)

        conn.commit()
        conn.close()
        return


    def fetch_from_db(mfg_pn, cycle, plant):
        # 读取数据库中的BOM，以result存储并返回
        conn = sqlite3.connect(os.path.join(CFE.path, CFE.filename))
        c = conn.cursor()
        cursor = c.execute(
            "SELECT * FROM {} WHERE MFG_PN='{}' AND Cycle='{}' AND Plant='{}'".format(CFE.bom_table, mfg_pn, cycle,
                                                                                  plant))

        result = cursor.fetchall()

        return result


    def del_bom_from_db(mfg_pn, cycle, plant):
        # 删除数据库中特定mfg_pn，特定cycle,特定plant 下的bom
        conn = sqlite3.connect(os.path.join(CFE.path, CFE.filename))
        c = conn.cursor()

        c.execute(
            "DELETE FROM {} WHERE MFG_PN='{}' AND Cycle='{}' AND Plant='{}'".format(CFE.bom_table, mfg_pn, cycle, plant))

        conn.commit()
        conn.close()

        return

    def failure_pn_into_db():
        # 将读取失败的pn、pn_desc、productfamily、cycle、plant信息写入数据库失效表
        conn = sqlite3.connect(os.path.join(CFE.path,CFE.filename))
        c = conn.cursor()
        c.execute('DELETE FROM {}'.format(CFE.failure_report_table))        # delete all records from failure report table
        place_holder = ','.join('?' * 5)  # make a {?,?,?,...,?}
        for row in CFE.failure_report:
            c.execute("INSERT INTO {} VALUES ({})".format(CFE.failure_report_table,place_holder),row)

        conn.commit()
        conn.close()
        return

    def write_into_excel(bom_array, pn, pn_desc, product_family, cycle, plant, excel_filename=None, path = r"D:\others\test"):
        # write bom_array into an excel file
        wb = Workbook()
        ws = wb.active
        ws.title = pn

        # write columns title
        for i in range(len(CFE.table_title)):
            ws.cell(column=i+1,row=1).value = CFE.table_title[i]
        # write bom data
        for row in range(len(bom_array)):
            for col in range(len(bom_array[row])):
                ws.cell(column=col+1,row=row+2).value = bom_array[row][col]
        
        if excel_filename==None:
            filename = os.path.join(path,' '.join((pn,pn_desc,cycle,plant,'.xlsx')))
        else:
            filename = os.path.join(path,excel_filename)
        wb.save(filename)

        return filename #return excel file name with absolute path

    def test2():
        pn='PA1R0003ID'
        pn_desc='Lenovo Phone A1000 ID 8G WH'
        productfamily='A1000'.upper()
        cycle='PH201701FINAL_B'
        plant='Ontim'.upper()
        
        
        bom_array=CFE.search_in_cfe(pn,pn_desc,productfamily,cycle,plant)
        for line in bom_array:
            print(line)
        print(len(bom_array))
        CFE.boms.put_nowait(StopIteration)
        CFE.write_into_excel(bom_array,pn,pn_desc,productfamily,cycle,plant)
        return

    def singlebom_export(path,pn,pn_desc,product_family, cycle,plant):
        # export one single MFG bom from CFE and write its bom into an excel file
        # path: the excel file path
        bom_array = CFE.search_in_cfe(pn,pn_desc,product_family,cycle,plant)
        #excel_filename = None
        if not bom_array == []:
            filename = CFE.write_into_excel(bom_array,pn,pn_desc,product_family,cycle,plant,path=path)
            return ('success', filename)
        else:
            return ('failure', None)


    def multibom_export_gevent(path,filename):
        # Use an excel file to record target boms, and export them from CFE
        # write these boms into seperate excels
        # the excel file containing all target boms has five columns: 
        # column A=pn, column B=pn description, column C=product family, column D=cycle, column E=plant
        # path: excel file path, and also the output boms excel path
        # filename: excel file name with only one sheet named "boms"
        # NO BLANK LINES! The first blank line means the list is ended.
        wb = load_workbook(os.path.join(path, filename))
        ws = wb.get_sheet_by_name('boms')
        pns = []
        pn_descs = []
        product_families = []
        cycles = []
        plants = []
        for loop in range(3):
            row = 2
            num_of_boms = 0
            pns[:] = []
            pn_descs[:] = []
            product_families[:] = []
            cycles[:] = []
            plants[:] = []
            while True:
                if (not ws.cell(column=1,row=row).value==None) and (not ws.cell(column=6, row=row).value=='success'):
                    pns.append(ws.cell(column=1,row=row).value)
                    pn_descs.append(ws.cell(column=2,row=row).value)
                    product_families.append(str(ws.cell(column=3,row=row).value).upper())
                    cycles.append(ws.cell(column=4,row=row).value)
                    plants.append(ws.cell(column=5,row=row).value)
                    row += 1
                else:
                    break

            print('bom list loaded, {} pns are to be exported'.format(len(pns)))
            num_of_boms = len(pns)
            # use greenlet to load bom from CFE
            gs = []
            for i in range(num_of_boms):
                g = Greenlet.spawn(CFE.singlebom_export,path,pns[i],pn_descs[i],product_families[i],cycles[i],plants[i])
                gs.append(g)
            gevent.joinall(gs)
            for i in range(num_of_boms):
                ws.cell(column=6,row=i+2).value=gs[i].value
            wb.save(os.path.join(path,filename))
            print("In {} loop, success bom number is {} while failure bom number is {}".format(loop,len(CFE.success_report), len(CFE.failure_report)))
            CFE.success_report[:] = []
            CFE.failure_report[:] = []
        return
    
    def multibom_export_loop(path,filename):
        # Use an excel file to record target boms, and export them from CFE
        # write these boms into seperate excels
        # the excel file containing all target boms has five columns: 
        # column A=pn, column B=pn description, column C=product family, column D=cycle, column E=plant
        # path: excel file path, and also the output boms excel path
        # filename: excel file name with only one sheet named "boms"
        # NO BLANK LINES! The first blank line means the list is ended.
        wb = load_workbook(os.path.join(path, filename))
        ws = wb.get_sheet_by_name('boms')
        pns = []
        pn_descs = []
        product_families = []
        cycles = []
        plants = []
        for loop in range(1):
            row = 2
            num_of_boms = 0
            pns[:] = []
            pn_descs[:] = []
            product_families[:] = []
            cycles[:] = []
            plants[:] = []
            while True:
                #print(row, ws.cell(column=6,row=row).value)
                if (not ws.cell(column=1,row=row).value==None):
                    if not ws.cell(column=6,row=row).value=='success':
                        """
                        pns.append(ws.cell(column=1,row=row).value)
                        pn_descs.append(ws.cell(column=2,row=row).value)
                        product_families.append(str(ws.cell(column=3,row=row).value).upper())
                        cycles.append(ws.cell(column=4,row=row).value)
                        plants.append(str(ws.cell(column=5,row=row).value))
                        """
                        pn = ws.cell(column=1,row=row).value
                        pn_desc = ws.cell(column=2,row=row).value
                        product_family = str(ws.cell(column=3,row=row).value).upper()
                        cycle = ws.cell(column=4,row=row).value
                        plant = str(ws.cell(column=5,row=row).value)
                        # status is 'success' or 'failure' returned by singlebom_export, indicating whether the bom is successfully exported from CFE
                        # filepath indicates the bom file location
                        (status, filepath) = CFE.singlebom_export(path, pn, pn_desc, product_family, cycle, plant)
                        ws.cell(column=6,row=row).value = status
                        if status == 'success':
                            ws.cell(column=7,row=row).value = '=HYPERLINK("{}")'.format(filepath)

                        #num_of_boms += 1
                    row += 1
                else:
                    break
            """
            print('bom list loaded, {} pns are to be exported'.format(len(pns)))
            num_of_boms = len(pns)
            for i in range(num_of_boms):
                # status is 'success' or 'failure' returned by singlebom_export, indicating whether the bom is successfully exported from CFE
                (status, filepath) = CFE.singlebom_export(path,pns[i],pn_descs[i],product_families[i],cycles[i],plants[i])
                print(pns[i],filepath,i)
                ws.cell(column=6,row=i+2).value = status
                if status == 'success':
                    ws.cell(column=7,row=i+2).value = '=HYPERLINK("{}")'.format(filepath)
            """
            
            print("In {} loop, success bom number is {} while failure bom number is {}".format(loop,len(CFE.success_report), len(CFE.failure_report)))
            CFE.success_report[:] = []
            CFE.failure_report[:] = []
        wb.save(os.path.join(path,filename))
        return
        """

    def multibom_export_loop(path,filename):
        # Use an excel file to record target boms, and export them from CFE
        # write these boms into seperate excels
        # the excel file containing all target boms has five columns: 
        # column A=pn, column B=pn description, column C=product family, column D=cycle, column E=plant
        # path: excel file path, and also the output boms excel path
        # filename: excel file name with only one sheet named "boms"
        # NO BLANK LINES! The first blank line means the list is ended.
        wb = load_workbook(os.path.join(path, filename))
        ws = wb.get_sheet_by_name('boms')
        
        row = 2
        pns = []
        pn_descs = []
        product_families = []
        cycles = []
        plants = []
        while True:
            if (not ws.cell(column=1,row=row).value==None) and (not ws.cell(column=6,row=row).value=='success') :
                pns.append(ws.cell(column=1,row=row).value)
                pn_descs.append(ws.cell(column=2,row=row).value)
                product_families.append(str(ws.cell(column=3,row=row).value).upper())
                cycles.append(ws.cell(column=4,row=row).value)
                plants.append(ws.cell(column=5,row=row).value)
                row += 1
            else:
                break
        num_of_boms = row - 2
        for i in range(num_of_boms):
            value = CFE.singlebom_export(path,pns[i],pn_descs[i],product_families[i],cycles[i],plants[i])
            ws.cell(column=6,row=i+2).value=value

        wb.save(os.path.join(path,filename))
        return
        """